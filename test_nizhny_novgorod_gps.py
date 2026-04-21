"""
Тест поиска суда по адресу в Нижнем Новгороде: Казанское шоссе 14к3.
Создаёт Excel с одной строкой (ФИО + адрес), запускает поиск по GPS/адресу и запись 45 полей.
Запуск из корня проекта:
  python -m tests.test_nizhny_novgorod_gps
  python tests/test_nizhny_novgorod_gps.py
"""
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pytest

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass


# Тестовые данные: ФИО из Нижнего Новгорода, адрес Казанское шоссе 14к3
TEST_FIO = "Петров Пётр Петрович"
TEST_ADDRESS = "Нижний Новгород, Казанское шоссе 14к3"
TEST_DEBT = 50_000  # руб., для расчёта госпошлины


def _check_yandex_and_db():
    """Проверяет геокодер Yandex и наличие судов по Нижегородской обл. в БД."""
    import os
    key = (os.getenv("YANDEX_GEO_KEY") or os.getenv("YANDEX_GEOCODER_API_KEY") or "").strip()
    yandex_ok = False
    if key:
        try:
            sys.path.insert(0, str(ROOT / "parser"))
            from check_apis import check_yandex_geocoder
            yandex_ok, msg = check_yandex_geocoder()
        except Exception:
            try:
                import requests
                r = requests.get(
                    "https://geocode-maps.yandex.ru/1.x/",
                    params={"apikey": key, "geocode": TEST_ADDRESS, "format": "json", "results": 1},
                    timeout=10,
                )
                yandex_ok = r.ok and bool(
                    r.json().get("response", {}).get("GeoObjectCollection", {}).get("featureMember", [])
                )
            except Exception:
                pass
    db_has_nnov = False
    total = 0
    courts_path = ROOT / "parser" / "courts.sqlite"
    if courts_path.exists():
        try:
            import sqlite3
            conn = sqlite3.connect(str(courts_path))
            cur = conn.execute(
                "SELECT COUNT(*) FROM courts WHERE LOWER(region) LIKE ?",
                ("%нижегород%",),
            )
            db_has_nnov = cur.fetchone()[0] > 0
            cur2 = conn.execute("SELECT COUNT(*) FROM courts")
            total = cur2.fetchone()[0]
            conn.close()
        except Exception:
            pass
    return yandex_ok, db_has_nnov, total


def test_nizhny_by_address():
    """Поиск суда по адресу и вывод 45 полей (без Excel)."""
    from court_locator.main import CourtLocator
    from court_locator.court_details import build_court_details, COURT_DETAIL_COLUMNS

    yandex_ok, db_has_nnov, total_courts = _check_yandex_and_db()
    if not yandex_ok:
        pytest.skip("Нет доступа к Yandex Geocoder (YANDEX_GEO_KEY/Locator в .env) или сервис недоступен.")
    if not db_has_nnov:
        pytest.skip("В `parser/courts.sqlite` нет судов по Нижегородской области — тест бессмысленен.")

    print("[OK] Yandex Geocoder доступен, в БД Нижний Новгород есть.")
    locator = CourtLocator(use_cache=True)
    try:
        court = locator.locate_court(address=TEST_ADDRESS)
        assert court, f"Суд не найден по адресу: {TEST_ADDRESS} при наличии ключа и судов в БД"
        print(f"[OK] Суд найден: {court.get('court_name', '')}")
        print(f"     Адрес суда: {court.get('address', '')}")
        print(f"     Регион: {court.get('region', '')}, участок: {court.get('section_num', '')}")

        details = build_court_details(court, normalized_address=TEST_ADDRESS, debt_amount=TEST_DEBT)
        print("\n--- 45 полей сведений ---")
        for col in COURT_DETAIL_COLUMNS:
            val = details.get(col, "")
            if val:
                print(f"  {col}: {val}")
        print("\nТест по адресу пройден.")
    finally:
        locator.close()


def test_nizhny_excel_pipeline():
    """Создаёт Excel с одной строкой (ФИО, Адрес, Сумма долга), запускает excel_court_by_gps, проверяет результат."""
    try:
        import openpyxl
    except ImportError:
        print("Пропуск теста Excel: pip install openpyxl")
        return

    from excel_court_by_gps import process_excel

    # Входной файл: одна строка
    test_dir = ROOT / "tests"
    test_dir.mkdir(exist_ok=True)
    input_xlsx = test_dir / "test_nizhny_input.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ФИО", "Адрес", "Сумма долга"])
    ws.append([TEST_FIO, TEST_ADDRESS, TEST_DEBT])
    wb.save(input_xlsx)
    wb.close()

    # Запуск обработки
    output_xlsx = test_dir / "test_nizhny_result.xlsx"
    out_path = process_excel(str(input_xlsx), str(output_xlsx))
    assert Path(out_path).exists(), f"Файл не создан: {out_path}"

    yandex_ok, db_has_nnov, _ = _check_yandex_and_db()
    if not yandex_ok:
        pytest.skip("Нет доступа к Yandex Geocoder (YANDEX_GEO_KEY/Locator в .env) или сервис недоступен.")
    if not db_has_nnov:
        pytest.skip("В `parser/courts.sqlite` нет судов по Нижегородской области — тест бессмысленен.")

    # Проверка: в результате есть колонки с данными суда
    wb2 = openpyxl.load_workbook(out_path)
    ws2 = wb2.active
    headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    assert "Наименование суда" in headers, "Колонка «Наименование суда» должна быть в результате"
    name_col = headers.index("Наименование суда") + 1
    court_name = ws2.cell(row=2, column=name_col).value
    wb2.close()

    print(f"[OK] Excel-пайплайн: результат записан в {out_path}")
    assert court_name and str(court_name).strip(), "Ожидали найденный суд для Нижнего Новгорода, но получили пустое значение"


if __name__ == "__main__":
    print("Тест: Нижний Новгород, Казанское шоссе 14к3")
    print("ФИО:", TEST_FIO)
    print("Адрес:", TEST_ADDRESS)
    print()
    test_nizhny_by_address()
    print()
    test_nizhny_excel_pipeline()
